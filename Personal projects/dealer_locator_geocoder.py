import requests
import time
import pandas as pd
import openpyxl as xl


def extract_cords(name, address, contact):
    var_dict = {'name': name,  'address': address, 'contact': contact}
    key = 'API_key'
    base_url = "https://maps.googleapis.com/maps/api/"
    place_url = f"place/textsearch/json?query={name}&key={key}&region=ZA"
    geocode_url = f"geocode/json?address={address}&key={key}&region=ZA"
    cords_list = []

    for item, value in var_dict.items():
        current_delay = 0.1
        max_delay = 10
        if item == 'name':
            url = base_url + place_url
        else:
            url = base_url + geocode_url

        while True:
            try:
                geo = requests.get(url)
            except IOError:
                pass
            else:
                if geo.status_code in range(200, 299):
                    try:
                        results = geo.json()['results'][0]
                        lat = results['geometry']['location']['lat']
                        lng = results['geometry']['location']['lng']

                        if lat < -34.83 or lat > -10 or lng < 16.49 or lng > 32.89:
                            cords_list.append('Wrong co-ordinates')
                            cords_list.append('Wrong co-ordinates')
                        else:
                            cords_list.append(lat)
                            cords_list.append(lng)
                        break
                    except IndexError:
                        pass
                elif geo.status_code in range(200, 299):
                    cords_list.append('Unsuccessful')
                    cords_list.append('Unsuccessful')
                    break

            if current_delay > max_delay:
                cords_list.append('Took too long')
                cords_list.append('Took too long')
                break
            time.sleep(current_delay)
            current_delay *= 2
    print(cords_list)
    return cords_list


def address_table(excel):
    wb = xl.load_workbook(excel)
    sheet = wb['Sheet1']
    add_list = []
    for row in range(2, sheet.max_row + 1):
        name = sheet.cell(row, 2)
        address = sheet.cell(row, 3)
        contact = sheet.cell(row, 4)
        name_co_ords = extract_cords(name.value, address.value, contact.value)
        info = name.value, address.value, contact.value, name_co_ords[0], name_co_ords[1], name_co_ords[2],\
            name_co_ords[3], name_co_ords[4], name_co_ords[5]
        add_list.append(info)
        df = pd.DataFrame(add_list, columns=['Name', 'Address', 'Contact', 'Name Lat', 'Name Long', 'Address Lat',
                                             'Address Long', 'Contact Lat', 'Contact Long'])
        writer = pd.ExcelWriter(r"filepath\filename.xlsx", engine='xlsxwriter')
        df.to_excel(writer, sheet_name='Sheet1', index=False)
        writer.close()
        print(info)

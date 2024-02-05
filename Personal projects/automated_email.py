import smtplib
import imaplib
import openpyxl as xl
import pandas as pd
from datetime import date
from datetime import datetime

sender_email = 'sender_email'
email_password = 'sender_password'
mail = imaplib.IMAP4_SSL('imap.gmail.com')
smtp_server = 'smtp.gmail.com'
smtp_port = 587
mail.login(sender_email, email_password)
smtp = smtplib.SMTP(smtp_server, smtp_port)

mailbox_name_to_select = b'"[Gmail]/Sent Mail"'
mail.select(mailbox_name_to_select)

smtp.starttls()
smtp.login(sender_email, email_password)
excel = "file path"
wb = xl.load_workbook(excel)
sheet = wb['Sheet1']


def sent_verification(dealer_email, search_date, sent_date_cell):
    sent_date = pd.to_datetime(sent_date_cell).date()
    to_email = f'{dealer_email}'
    search_criteria = f'(TO "{dealer_email}" SENTSINCE {search_date})'
    status, messages = mail.search(None, search_criteria)
    email_ids = messages[0].split()
    return bool(email_ids)

def email_reminder(dealer_email):
    subject = 'Subject'
    current_time = datetime.now().time()
    current_hour = current_time.hour
    if 0 <= current_hour < 12:
        greeting = "Good morning"
    else:
        greeting = "Good afternoon"
    body = f'{greeting},\n\nAccording to our records your licence expired on {expiry_date}.' \
           f' Kindly send the renewed licence so that we may update our records.\n'
    message = f'Subject: {subject}\n\n{body}'
    smtp.sendmail(from_email, dealer_email, message)


for row in range(2, 4):
    expiry_date = sheet.cell(row, 8).value
    sent_date_cell = sheet.cell(row, 11).value

    while expiry_date is not None:
        expiry_date = pd.to_datetime(expiry_date).date()
        expiry_duration = date.today() - expiry_date
        dealer_email = sheet.cell(row, 7).value
        pop_payment = sheet.cell(row, 9).value
        sent_request = sheet.cell(row, 10).value

        if sent_date_cell is None and expiry_duration.days > 30:
            email_reminder(dealer_email)
            sheet[f'J{row}'] = 'Yes'
            sheet[f'K{row}'] = f'{date.today()}'
            wb.save(excel)
            break

        search_date = pd.to_datetime(sent_date_cell).strftime('%d-%b-%Y')
        if sent_request == "Yes" and sent_verification(dealer_email, search_date, sent_date_cell) == True and expiry_duration > 30:
            email_reminder(dealer_email)
            break

        elif sent_request == "Yes" and sent_verification(dealer_email, search_date, sent_date_cell) == False:
            email_reminder(dealer_email)
            sheet[f'J{row}'] = 'Yes'
            wb.save(excel)
            break

        elif sent_request == "Not yet" and expiry_duration < 0:
            break

smtp.quit()
mail.logout()

